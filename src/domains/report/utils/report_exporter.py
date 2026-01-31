"""
Report Export Utility - Generate downloadable reports
Developed By Samuel Kufre Willie - 31 January 2026
"""

import io
import csv
from datetime import datetime
from typing import Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER


class ReportExporter:
    """Handle export of analytics reports in various formats"""

    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()

    def _setup_custom_styles(self):
        """Setup custom PDF styles"""
        self.styles.add(
            ParagraphStyle(
                name="CustomTitle",
                parent=self.styles["Heading1"],
                fontSize=24,
                textColor=colors.HexColor("#BF4C20"),  # Kidemia primary
                spaceAfter=30,
                alignment=TA_CENTER,
            )
        )
        self.styles.add(
            ParagraphStyle(
                name="SectionHeading",
                parent=self.styles["Heading2"],
                fontSize=16,
                textColor=colors.HexColor("#6366F1"),  # Kidemia secondary
                spaceAfter=12,
                spaceBefore=12,
            )
        )

    # ==================== CSV EXPORT ====================

    def export_to_csv(self, data: Dict[str, Any], report_type: str) -> io.StringIO:
        """Export data to CSV format"""
        output = io.StringIO()
        writer = csv.writer(output)

        # Write header
        writer.writerow([f"Kidemia Analytics Report - {report_type}"])
        writer.writerow([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        writer.writerow([])  # Empty row

        if report_type == "platform_overview":
            self._write_platform_overview_csv(writer, data)
        elif report_type == "student_performance":
            self._write_student_performance_csv(writer, data)
        elif report_type == "financial":
            self._write_financial_csv(writer, data)
        elif report_type == "assessment_analysis":
            self._write_assessment_csv(writer, data)
        elif report_type == "question_quality":
            self._write_question_quality_csv(writer, data)

        output.seek(0)
        return output

    def _write_platform_overview_csv(self, writer, data):
        """Write platform overview data to CSV"""
        overview = data.get("overview", {})

        writer.writerow(["Platform Statistics"])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Total Users", overview.get("total_users", 0)])
        writer.writerow(["Total Students", overview.get("total_students", 0)])
        writer.writerow(["Total Assessments", overview.get("total_assessments", 0)])
        writer.writerow(
            ["Active Subscriptions", overview.get("active_subscriptions", 0)]
        )
        writer.writerow(["Total Revenue", f"₦{overview.get('total_revenue', 0):,.2f}"])
        writer.writerow(["Completion Rate", f"{overview.get('completion_rate', 0)}%"])
        writer.writerow([])

        # Revenue data
        revenue = data.get("revenue", {})
        writer.writerow(["Revenue Breakdown"])
        writer.writerow(["Category", "Amount"])
        writer.writerow(
            ["Subscriptions", f"₦{revenue.get('subscription_revenue', 0):,.2f}"]
        )
        writer.writerow(
            ["Assessments", f"₦{revenue.get('assessment_revenue', 0):,.2f}"]
        )
        writer.writerow(["Wallet Top-ups", f"₦{revenue.get('wallet_topup', 0):,.2f}"])

    def _write_student_performance_csv(self, writer, data):
        """Write student performance data to CSV"""
        perf = data.get("performance_summary", {})

        writer.writerow(["Student Performance Summary"])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Total Attempts", perf.get("total_attempts", 0)])
        writer.writerow(["Completed", perf.get("completed_attempts", 0)])
        writer.writerow(["Average Score", f"{perf.get('average_score', 0)}%"])
        writer.writerow(["Pass Rate", f"{perf.get('pass_rate', 0)}%"])
        writer.writerow(["Best Score", f"{perf.get('best_score', 0)}%"])
        writer.writerow([])

        # Subject breakdown
        subjects = data.get("subject_breakdown", [])
        if subjects:
            writer.writerow(["Subject Performance"])
            writer.writerow(["Subject", "Attempts", "Average Score", "Pass Rate"])
            for subj in subjects:
                writer.writerow(
                    [
                        subj["subject_name"],
                        subj["total_attempts"],
                        f"{subj['average_score']}%",
                        f"{subj['pass_rate']}%",
                    ]
                )

        writer.writerow([])

        # Topic breakdown (NEW)
        topics = data.get("topic_breakdown", [])
        if topics:
            writer.writerow(["Topic Performance"])
            writer.writerow(
                ["Subject", "Topic", "Questions", "Success Rate", "Mastery Level"]
            )
            for topic in topics:
                writer.writerow(
                    [
                        topic["subject_name"],
                        topic["topic_name"],
                        topic["questions_attempted"],
                        f"{topic['success_rate']}%",
                        topic["mastery_level"],
                    ]
                )

    def _write_financial_csv(self, writer, data):
        """Write financial data to CSV"""
        overview = data.get("overview", {})

        writer.writerow(["Financial Overview"])
        writer.writerow(["Metric", "Amount"])
        writer.writerow(["Total Revenue", f"₦{overview.get('total_revenue', 0):,.2f}"])
        writer.writerow(
            ["Monthly Revenue", f"₦{overview.get('monthly_revenue', 0):,.2f}"]
        )
        writer.writerow(
            ["Avg Transaction", f"₦{overview.get('average_transaction_value', 0):,.2f}"]
        )
        writer.writerow([])

        # Revenue trend
        trend = data.get("trend", [])
        if trend:
            writer.writerow(["Daily Revenue Trend"])
            writer.writerow(["Date", "Revenue", "Transactions"])
            for item in trend:
                writer.writerow(
                    [item["date"], f"₦{item['revenue']:,.2f}", item["transactions"]]
                )

    def _write_assessment_csv(self, writer, data):
        """Write assessment analytics to CSV"""
        report = data.get("report", {})
        assessment_info = report.get("assessment", {})
        attempts = report.get("attempts", {})
        scores = report.get("scores", {})

        writer.writerow(["Assessment Analytics"])
        writer.writerow(["Assessment", assessment_info.get("title", "N/A")])
        writer.writerow(["Category", assessment_info.get("category", "N/A")])
        writer.writerow([])

        writer.writerow(["Attempt Statistics"])
        writer.writerow(["Total Attempts", attempts.get("total", 0)])
        writer.writerow(["Completed", attempts.get("completed", 0)])
        writer.writerow(["Passed", attempts.get("passed", 0)])
        writer.writerow(["Pass Rate", f"{attempts.get('pass_rate', 0)}%"])
        writer.writerow([])

        writer.writerow(["Score Statistics"])
        writer.writerow(["Average", f"{scores.get('average', 0)}%"])
        writer.writerow(["Minimum", f"{scores.get('minimum', 0)}%"])
        writer.writerow(["Maximum", f"{scores.get('maximum', 0)}%"])

    def _write_question_quality_csv(self, writer, data):
        """Write question quality data to CSV"""
        writer.writerow(["Question Quality Analysis"])
        writer.writerow(["Total Analyzed", data.get("total_analyzed", 0)])
        writer.writerow(
            ["Needs Adjustment", data.get("needs_difficulty_adjustment", 0)]
        )
        writer.writerow([])

        questions = data.get("most_missed_questions", [])
        if questions:
            writer.writerow(["Most Missed Questions"])
            writer.writerow(
                ["Question Preview", "Subject", "Difficulty", "Success Rate"]
            )
            for q in questions:
                writer.writerow(
                    [
                        q["question_preview"],
                        q["subject"],
                        q["difficulty"],
                        f"{q['success_rate']}%",
                    ]
                )

    def export_to_excel(self, data: Dict[str, Any], report_type: str) -> io.BytesIO:
        """Export data to Excel format with formatting"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        if report_type == "platform_overview":
            self._create_platform_overview_excel(wb, data)
        elif report_type == "student_performance":
            self._create_student_performance_excel(wb, data)
        elif report_type == "financial":
            self._create_financial_excel(wb, data)
        elif report_type == "assessment_analysis":
            self._create_assessment_excel(wb, data)
        elif report_type == "question_quality":
            self._create_question_quality_excel(wb, data)

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _create_platform_overview_excel(self, wb, data):
        """Create platform overview Excel sheets"""
        # Overview sheet
        ws = wb.create_sheet("Overview")
        overview = data.get("overview", {})

        # Header
        ws["A1"] = "Kidemia Platform Analytics"
        ws["A1"].font = Font(size=18, bold=True, color="BF4C20")
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # Data
        row = 4
        ws[f"A{row}"] = "Metric"
        ws[f"B{row}"] = "Value"
        self._style_header_row(ws, row)

        metrics = [
            ("Total Users", overview.get("total_users", 0)),
            ("Total Students", overview.get("total_students", 0)),
            ("Total Assessments", overview.get("total_assessments", 0)),
            ("Active Subscriptions", overview.get("active_subscriptions", 0)),
            ("Total Revenue", f"₦{overview.get('total_revenue', 0):,.2f}"),
            ("Completion Rate", f"{overview.get('completion_rate', 0)}%"),
        ]

        for metric, value in metrics:
            row += 1
            ws[f"A{row}"] = metric
            ws[f"B{row}"] = value

        self._auto_adjust_columns(ws)

    def _create_student_performance_excel(self, wb, data):
        """Create student performance Excel sheets"""
        # Summary sheet
        ws_summary = wb.create_sheet("Performance Summary")
        perf = data.get("performance_summary", {})

        ws_summary["A1"] = "Student Performance Report"
        ws_summary["A1"].font = Font(size=18, bold=True, color="BF4C20")

        row = 3
        ws_summary[f"A{row}"] = "Metric"
        ws_summary[f"B{row}"] = "Value"
        self._style_header_row(ws_summary, row)

        metrics = [
            ("Total Attempts", perf.get("total_attempts", 0)),
            ("Completed", perf.get("completed_attempts", 0)),
            ("Average Score", f"{perf.get('average_score', 0)}%"),
            ("Pass Rate", f"{perf.get('pass_rate', 0)}%"),
            ("Best Score", f"{perf.get('best_score', 0)}%"),
            ("Worst Score", f"{perf.get('worst_score', 0)}%"),
        ]

        for metric, value in metrics:
            row += 1
            ws_summary[f"A{row}"] = metric
            ws_summary[f"B{row}"] = value

        # Subject performance sheet
        ws_subjects = wb.create_sheet("Subject Performance")
        ws_subjects["A1"] = "Subject Breakdown"
        ws_subjects["A1"].font = Font(size=16, bold=True, color="6366F1")

        subjects = data.get("subject_breakdown", [])
        if subjects:
            headers = ["Subject", "Attempts", "Avg Score", "Pass Rate"]
            for col, header in enumerate(headers, 1):
                ws_subjects.cell(3, col, header)
                self._style_cell(ws_subjects.cell(3, col), is_header=True)

            for row_idx, subj in enumerate(subjects, 4):
                ws_subjects.cell(row_idx, 1, subj["subject_name"])
                ws_subjects.cell(row_idx, 2, subj["total_attempts"])
                ws_subjects.cell(row_idx, 3, f"{subj['average_score']}%")
                ws_subjects.cell(row_idx, 4, f"{subj['pass_rate']}%")

        # Topic performance sheet (NEW)
        ws_topics = wb.create_sheet("Topic Performance")
        ws_topics["A1"] = "Topic-Level Analysis"
        ws_topics["A1"].font = Font(size=16, bold=True, color="6366F1")

        topics = data.get("topic_breakdown", [])
        if topics:
            headers = [
                "Subject",
                "Topic",
                "Questions",
                "Correct",
                "Success Rate",
                "Mastery",
            ]
            for col, header in enumerate(headers, 1):
                ws_topics.cell(3, col, header)
                self._style_cell(ws_topics.cell(3, col), is_header=True)

            for row_idx, topic in enumerate(topics, 4):
                ws_topics.cell(row_idx, 1, topic["subject_name"])
                ws_topics.cell(row_idx, 2, topic["topic_name"])
                ws_topics.cell(row_idx, 3, topic["questions_attempted"])
                ws_topics.cell(row_idx, 4, topic["correct_answers"])
                ws_topics.cell(row_idx, 5, f"{topic['success_rate']}%")
                ws_topics.cell(row_idx, 6, topic["mastery_level"])

                # Color code mastery level
                mastery_cell = ws_topics.cell(row_idx, 6)
                if topic["mastery_level"] == "MASTERED":
                    mastery_cell.fill = PatternFill(
                        start_color="10B981", fill_type="solid"
                    )
                elif topic["mastery_level"] == "NEEDS_WORK":
                    mastery_cell.fill = PatternFill(
                        start_color="EF4444", fill_type="solid"
                    )

        for ws in [ws_summary, ws_subjects, ws_topics]:
            self._auto_adjust_columns(ws)

    def _create_financial_excel(self, wb, data):
        """Create financial Excel sheets"""
        ws = wb.create_sheet("Financial Overview")
        overview = data.get("overview", {})

        ws["A1"] = "Financial Report"
        ws["A1"].font = Font(size=18, bold=True, color="BF4C20")

        row = 3
        ws[f"A{row}"] = "Metric"
        ws[f"B{row}"] = "Amount (₦)"
        self._style_header_row(ws, row)

        metrics = [
            ("Total Revenue", overview.get("total_revenue", 0)),
            ("Monthly Revenue", overview.get("monthly_revenue", 0)),
            ("Subscription Revenue", overview.get("subscription_revenue", 0)),
            ("Assessment Revenue", overview.get("assessment_revenue", 0)),
            ("Wallet Top-ups", overview.get("wallet_topup", 0)),
        ]

        for metric, value in metrics:
            row += 1
            ws[f"A{row}"] = metric
            ws[f"B{row}"] = f"₦{value:,.2f}"

        # Revenue trend sheet
        ws_trend = wb.create_sheet("Revenue Trend")
        trend = data.get("trend", [])
        if trend:
            ws_trend["A1"] = "Daily Revenue"
            ws_trend["A1"].font = Font(size=16, bold=True)

            headers = ["Date", "Revenue", "Transactions"]
            for col, header in enumerate(headers, 1):
                ws_trend.cell(3, col, header)
                self._style_cell(ws_trend.cell(3, col), is_header=True)

            for row_idx, item in enumerate(trend, 4):
                ws_trend.cell(row_idx, 1, item["date"])
                ws_trend.cell(row_idx, 2, f"₦{item['revenue']:,.2f}")
                ws_trend.cell(row_idx, 3, item["transactions"])

        for ws in [ws, ws_trend]:
            self._auto_adjust_columns(ws)

    def _create_assessment_excel(self, wb, data):
        """Create assessment analytics Excel"""
        ws = wb.create_sheet("Assessment Report")
        report = data.get("report", {})
        assessment = report.get("assessment", {})

        ws["A1"] = f"Assessment: {assessment.get('title', 'N/A')}"
        ws["A1"].font = Font(size=18, bold=True, color="BF4C20")
        ws["A2"] = f"Category: {assessment.get('category', 'N/A')}"

        # Statistics
        attempts = report.get("attempts", {})
        scores = report.get("scores", {})

        row = 4
        sections = [
            (
                "Attempts",
                [
                    ("Total", attempts.get("total", 0)),
                    ("Completed", attempts.get("completed", 0)),
                    ("Passed", attempts.get("passed", 0)),
                    ("Pass Rate", f"{attempts.get('pass_rate', 0)}%"),
                ],
            ),
            (
                "Scores",
                [
                    ("Average", f"{scores.get('average', 0)}%"),
                    ("Minimum", f"{scores.get('minimum', 0)}%"),
                    ("Maximum", f"{scores.get('maximum', 0)}%"),
                    ("Std Deviation", f"{scores.get('standard_deviation', 0)}"),
                ],
            ),
        ]

        for section_name, section_data in sections:
            ws[f"A{row}"] = section_name
            ws[f"A{row}"].font = Font(bold=True, size=14, color="6366F1")
            row += 1

            for metric, value in section_data:
                ws[f"A{row}"] = metric
                ws[f"B{row}"] = value
                row += 1

            row += 1

        self._auto_adjust_columns(ws)

    def _create_question_quality_excel(self, wb, data):
        """Create question quality Excel"""
        ws = wb.create_sheet("Question Quality")

        ws["A1"] = "Question Quality Analysis"
        ws["A1"].font = Font(size=18, bold=True, color="BF4C20")

        ws["A3"] = "Total Analyzed"
        ws["B3"] = data.get("total_analyzed", 0)
        ws["A4"] = "Needs Adjustment"
        ws["B4"] = data.get("needs_difficulty_adjustment", 0)

        # Most missed questions
        questions = data.get("most_missed_questions", [])
        if questions:
            row = 6
            ws[f"A{row}"] = "Most Missed Questions"
            ws[f"A{row}"].font = Font(bold=True, size=14)

            row += 1
            headers = ["Question", "Subject", "Difficulty", "Success Rate"]
            for col, header in enumerate(headers, 1):
                ws.cell(row, col, header)
                self._style_cell(ws.cell(row, col), is_header=True)

            for q in questions:
                row += 1
                ws.cell(row, 1, q["question_preview"])
                ws.cell(row, 2, q["subject"])
                ws.cell(row, 3, q["difficulty"])
                ws.cell(row, 4, f"{q['success_rate']}%")

        self._auto_adjust_columns(ws)

    def _style_header_row(self, ws, row):
        """Apply header styling to a row"""
        for cell in ws[row]:
            self._style_cell(cell, is_header=True)

    def _style_cell(self, cell, is_header=False):
        """Apply consistent cell styling"""
        if is_header:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="6366F1", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        cell.border = thin_border

    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def export_to_pdf(self, data: Dict[str, Any], report_type: str) -> io.BytesIO:
        """Export data to PDF format"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []

        # Title
        title = Paragraph(
            f"Kidemia Analytics Report<br/><font size=12>{report_type.replace('_', ' ').title()}</font>",
            self.styles["CustomTitle"],
        )
        story.append(title)
        story.append(Spacer(1, 0.2 * inch))

        # Timestamp
        timestamp = Paragraph(
            f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
            self.styles["Normal"],
        )
        story.append(timestamp)
        story.append(Spacer(1, 0.3 * inch))

        if report_type == "platform_overview":
            self._add_platform_overview_pdf(story, data)
        elif report_type == "student_performance":
            self._add_student_performance_pdf(story, data)
        elif report_type == "financial":
            self._add_financial_pdf(story, data)
        elif report_type == "assessment_analysis":
            self._add_assessment_pdf(story, data)
        elif report_type == "question_quality":
            self._add_question_quality_pdf(story, data)

        doc.build(story)
        buffer.seek(0)
        return buffer

    def _add_platform_overview_pdf(self, story, data):
        """Add platform overview to PDF"""
        story.append(Paragraph("Platform Statistics", self.styles["SectionHeading"]))

        overview = data.get("overview", {})
        data_table = [
            ["Metric", "Value"],
            ["Total Users", f"{overview.get('total_users', 0):,}"],
            ["Total Students", f"{overview.get('total_students', 0):,}"],
            ["Total Assessments", f"{overview.get('total_assessments', 0):,}"],
            [
                "Active Subscriptions",
                f"{overview.get('active_subscriptions', 0):,}",
            ],
            ["Total Revenue", f"₦{overview.get('total_revenue', 0):,.2f}"],
            ["Completion Rate", f"{overview.get('completion_rate', 0)}%"],
        ]

        table = Table(data_table)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6366F1")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 0.3 * inch))

        # Revenue breakdown
        story.append(Paragraph("Revenue Breakdown", self.styles["SectionHeading"]))
        revenue = data.get("revenue", {})
        revenue_table = [
            ["Category", "Amount"],
            [
                "Subscriptions",
                f"₦{revenue.get('subscription_revenue', 0):,.2f}",
            ],
            [
                "Assessments",
                f"₦{revenue.get('assessment_revenue', 0):,.2f}",
            ],
            ["Wallet Top-ups", f"₦{revenue.get('wallet_topup', 0):,.2f}"],
        ]

        table = Table(revenue_table)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#BF4C20")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        story.append(table)

    def _add_student_performance_pdf(self, story, data):
        """Add student performance to PDF"""
        story.append(Paragraph("Performance Summary", self.styles["SectionHeading"]))

        perf = data.get("performance_summary", {})
        perf_table = [
            ["Metric", "Value"],
            ["Total Attempts", f"{perf.get('total_attempts', 0):,}"],
            ["Completed", f"{perf.get('completed_attempts', 0):,}"],
            ["Average Score", f"{perf.get('average_score', 0)}%"],
            ["Pass Rate", f"{perf.get('pass_rate', 0)}%"],
            ["Best Score", f"{perf.get('best_score', 0)}%"],
        ]

        table = Table(perf_table)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6366F1")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 0.3 * inch))

        # Subject breakdown
        subjects = data.get("subject_breakdown", [])
        if subjects:
            story.append(
                Paragraph("Subject Performance", self.styles["SectionHeading"])
            )
            subj_data = [["Subject", "Attempts", "Avg Score", "Pass Rate"]]
            for subj in subjects[:10]:  # Limit to 10 for PDF
                subj_data.append(
                    [
                        subj["subject_name"],
                        str(subj["total_attempts"]),
                        f"{subj['average_score']}%",
                        f"{subj['pass_rate']}%",
                    ]
                )

            table = Table(subj_data)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#10B981")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ]
                )
            )
            story.append(table)

    def _add_financial_pdf(self, story, data):
        """Add financial data to PDF"""
        story.append(Paragraph("Financial Overview", self.styles["SectionHeading"]))

        overview = data.get("overview", {})
        fin_table = [
            ["Metric", "Amount"],
            ["Total Revenue", f"₦{overview.get('total_revenue', 0):,.2f}"],
            ["Monthly Revenue", f"₦{overview.get('monthly_revenue', 0):,.2f}"],
            [
                "Subscription Revenue",
                f"₦{overview.get('subscription_revenue', 0):,.2f}",
            ],
            [
                "Assessment Revenue",
                f"₦{overview.get('assessment_revenue', 0):,.2f}",
            ],
            ["Total Transactions", f"{overview.get('total_transactions', 0):,}"],
        ]

        table = Table(fin_table)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F59E0B")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        story.append(table)

    def _add_assessment_pdf(self, story, data):
        """Add assessment analytics to PDF"""
        report = data.get("report", {})
        assessment = report.get("assessment", {})

        story.append(
            Paragraph(
                f"Assessment: {assessment.get('title', 'N/A')}",
                self.styles["SectionHeading"],
            )
        )

        attempts = report.get("attempts", {})
        scores = report.get("scores", {})

        stats_table = [
            ["Metric", "Value"],
            ["Total Attempts", f"{attempts.get('total', 0):,}"],
            ["Completed", f"{attempts.get('completed', 0):,}"],
            ["Pass Rate", f"{attempts.get('pass_rate', 0)}%"],
            ["Average Score", f"{scores.get('average', 0)}%"],
            ["Minimum Score", f"{scores.get('minimum', 0)}%"],
            ["Maximum Score", f"{scores.get('maximum', 0)}%"],
        ]

        table = Table(stats_table)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8B5CF6")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        story.append(table)

    def _add_question_quality_pdf(self, story, data):
        """Add question quality to PDF"""
        story.append(
            Paragraph("Question Quality Analysis", self.styles["SectionHeading"])
        )

        summary_table = [
            ["Metric", "Value"],
            ["Total Analyzed", f"{data.get('total_analyzed', 0):,}"],
            [
                "Needs Adjustment",
                f"{data.get('needs_difficulty_adjustment', 0):,}",
            ],
        ]

        table = Table(summary_table)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EC4899")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 0.2 * inch))

        # Most missed questions
        questions = data.get("most_missed_questions", [])
        if questions:
            story.append(
                Paragraph("Most Missed Questions", self.styles["SectionHeading"])
            )
            q_data = [["Question", "Subject", "Success Rate"]]
            for q in questions[:5]:  # Top 5 for PDF
                q_data.append(
                    [
                        q["question_preview"][:50] + "...",
                        q["subject"],
                        f"{q['success_rate']}%",
                    ]
                )

            table = Table(q_data, colWidths=[3 * inch, 1.5 * inch, 1 * inch])
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EF4444")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            story.append(table)
